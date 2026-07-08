"""
PUML Test Case Generator — Python Backend
==========================================
Supports any OpenAI-compatible local LLM:
  Ollama      → http://localhost:11434/v1
  LM Studio   → http://localhost:1234/v1
  Jan         → http://localhost:1337/v1

Usage:
    pip install flask flask-cors openpyxl requests
    python backend.py

    # Override LLM:
    LLM_BASE_URL=http://localhost:1234/v1 LLM_MODEL=mistral python backend.py
"""

import json, os, re, traceback, subprocess, tempfile, shutil, base64, hashlib
from io import BytesIO

import requests
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import os as _os
app = Flask(__name__, static_folder=_os.path.dirname(_os.path.abspath(__file__)))
CORS(app)

# ── PlantUML Rendering Configuration ─────────────────────────────────────────
# Place plantuml.jar in the same folder as this file (or set PLANTUML_JAR env var)
# to render diagrams locally and fully offline. Download it from:
#   https://plantuml.com/download
#
# If plantuml.jar is not found, rendering automatically falls back to the
# public Kroki.io service (requires internet access).
PLANTUML_JAR = _os.getenv(
    "PLANTUML_JAR",
    _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "plantuml.jar"),
)
JAVA_BIN     = _os.getenv("JAVA_BIN", "java")
RENDER_CACHE_DIR = _os.path.join(tempfile.gettempdir(), "puml_render_cache")
_os.makedirs(RENDER_CACHE_DIR, exist_ok=True)


def plantuml_jar_available():
    """Check whether plantuml.jar + Java are both usable."""
    if not _os.path.exists(PLANTUML_JAR):
        return False
    if shutil.which(JAVA_BIN) is None and not _os.path.isfile(JAVA_BIN):
        return False
    return True


def render_with_jar(puml_text: str, fmt: str = "png") -> bytes:
    """
    Render PUML to PNG or SVG bytes using local plantuml.jar via Java.
    fmt: 'png' or 'svg'
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        src_path = _os.path.join(tmpdir, "diagram.puml")
        with open(src_path, "w", encoding="utf-8") as f:
            f.write(puml_text)

        flag = "-tsvg" if fmt == "svg" else "-tpng"
        cmd = [JAVA_BIN, "-jar", PLANTUML_JAR, flag, "-o", tmpdir, src_path]

        result = subprocess.run(cmd, capture_output=True, timeout=30)
        if result.returncode != 0:
            err = result.stderr.decode("utf-8", errors="ignore")
            raise RuntimeError(f"plantuml.jar render failed: {err[:300]}")

        out_ext = "svg" if fmt == "svg" else "png"
        out_path = _os.path.join(tmpdir, f"diagram.{out_ext}")
        if not _os.path.exists(out_path):
            # plantuml sometimes names output after @startuml title
            candidates = [f for f in _os.listdir(tmpdir) if f.endswith(f".{out_ext}")]
            if not candidates:
                raise RuntimeError("plantuml.jar produced no output file")
            out_path = _os.path.join(tmpdir, candidates[0])

        with open(out_path, "rb") as f:
            return f.read()


def render_with_kroki(puml_text: str, fmt: str = "png") -> bytes:
    """Fallback renderer using the public Kroki.io API (requires internet)."""
    import zlib
    compressed = zlib.compress(puml_text.encode("utf-8"), level=9)
    encoded = base64.urlsafe_b64encode(compressed).decode("ascii")
    url = f"https://kroki.io/plantuml/{fmt}/{encoded}"
    resp = requests.get(url, timeout=20)
    resp.raise_for_status()
    return resp.content


def render_puml(puml_text: str, fmt: str = "png") -> bytes:
    """
    Render PUML text to an image. Tries local plantuml.jar first,
    falls back to Kroki.io if the jar is unavailable.
    """
    cache_key = hashlib.sha256((puml_text + fmt).encode("utf-8")).hexdigest()
    cache_path = _os.path.join(RENDER_CACHE_DIR, f"{cache_key}.{fmt}")
    if _os.path.exists(cache_path):
        with open(cache_path, "rb") as f:
            return f.read()

    if plantuml_jar_available():
        data = render_with_jar(puml_text, fmt)
    else:
        data = render_with_kroki(puml_text, fmt)

    with open(cache_path, "wb") as f:
        f.write(data)
    return data


@app.route("/")
def index():
    """Serve the frontend — open http://localhost:5050 in browser."""
    html_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "index.html")
    if _os.path.exists(html_path):
        return send_file(html_path)
    return "<h2>index.html not found — place it in the same folder as backend.py</h2>", 404

@app.route("/css/<path:filename>")
def serve_css(filename):
    css_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "css")
    return send_file(_os.path.join(css_dir, filename))

@app.route("/js/<path:filename>")
def serve_js(filename):
    js_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "js")
    return send_file(_os.path.join(js_dir, filename))

LLM_BASE_URL = os.getenv("LLM_BASE_URL", "http://localhost:11434/v1")
LLM_MODEL    = os.getenv("LLM_MODEL",    "llama3")
LLM_TIMEOUT  = int(os.getenv("LLM_TIMEOUT", "120"))


# ── PUML Parser ──────────────────────────────────────────────────────────────

def parse_puml(content):
    title_m = re.search(r"title\s+(.+)", content, re.IGNORECASE)
    title   = title_m.group(1).strip() if title_m else "System Flow"

    actors = []
    for m in re.finditer(r"(?:actor|participant|boundary|control|entity|database|component)\s+\"?([^\"\n]+)\"?(?:\s+as\s+(\w+))?", content, re.IGNORECASE):
        actors.append(m.group(2) or m.group(1).strip())

    messages = []
    for m in re.finditer(r"(\w+)\s*(?:->|-->|->>|-->>)\s*(\w+)\s*:\s*(.+)", content):
        messages.append({"from": m.group(1), "to": m.group(2), "label": m.group(3).strip()})

    conditions = []
    for m in re.finditer(r"(?:alt|else|opt|loop|break)\s+(.+)", content, re.IGNORECASE):
        conditions.append(m.group(1).strip())

    notes = []
    for m in re.finditer(r"note\s+(?:over|left|right)?\s*[^:]*:\s*(.+)", content, re.IGNORECASE):
        notes.append(m.group(1).strip())

    return {"title": title, "actors": actors, "messages": messages, "conditions": conditions, "notes": notes}


# ── LLM Prompt ───────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """You are a senior QA engineer writing a manual test case document.
Generate test cases that a tester will execute BY HAND — no execution results, no status.

Each test case MUST have EXACTLY these fields:
  tc_id          : "TC-001", "TC-002", etc.
  module         : short module or feature name extracted from the diagram
  name           : concise test case title
  type           : one of Positive / Negative / Functional / Security
  priority       : one of High / Medium / Low
  prerequisite   : conditions that must be true BEFORE the tester starts
                   (system state, data, accounts, config, dependencies)
  steps          : numbered list of exact manual actions the tester performs
                   written as step-by-step instructions
  expected_output: what the tester should observe / verify after executing steps

Rules:
- DO NOT include: status, actual result, pass/fail, comments, execution fields
- prerequisite must be concrete (e.g. "User account exists with role Admin. System is logged out.")
- steps must be actionable (e.g. "1. Open browser\n2. Navigate to /login\n3. Enter username...")
- expected_output must be verifiable (e.g. "System returns 200 OK with JSON array of user objects")
- Cover: happy path, negative/error cases, boundary values, auth/security, edge cases
- Minimum 8 test cases. Return ONLY a valid JSON array. No markdown, no explanation."""

def build_prompt(puml, parsed, knowledge=None):
    knowledge = knowledge or {}
    msgs = "\n".join(f"  {m['from']} -> {m['to']}: {m['label']}" for m in parsed["messages"])

    ctx_lines = []
    if knowledge.get("project"):     ctx_lines.append(f"Project: {knowledge['project']}")
    if knowledge.get("description"): ctx_lines.append(f"System Description: {knowledge['description']}")
    if knowledge.get("tech"):         ctx_lines.append(f"Tech Stack: {knowledge['tech']}")
    if knowledge.get("keywords"):     ctx_lines.append(f"Key Entities: {', '.join(knowledge['keywords'])}")
    if knowledge.get("rules"):        ctx_lines.append(f"Business Rules: {knowledge['rules']}")
    if knowledge.get("extra"):        ctx_lines.append(f"Additional Instructions: {knowledge['extra']}")

    focus = knowledge.get("focus", "all")
    focus_txt = {
        "positive": "Focus ONLY on positive / happy-path test cases.",
        "negative": "Focus ONLY on negative / edge-case test cases.",
        "security": "Focus heavily on security-related test cases (auth, tokens, access control).",
        "all":      "Cover positive, negative, and security test cases.",
    }.get(focus, "Cover positive, negative, and security test cases.")

    min_cases = knowledge.get("min_cases", 8)

    context_block = ""
    if ctx_lines:
        context_block = "\n=== KNOWLEDGE CONTEXT ===\n" + "\n".join(ctx_lines) + "\n"

    return (
        f"Analyze this PlantUML diagram and generate manual test cases.\n\n"
        f"Title: {parsed['title']}\n"
        f"Actors: {', '.join(parsed['actors']) or 'none'}\n"
        f"Messages:\n{msgs or '  (none)'}\n"
        f"Conditions: {', '.join(parsed['conditions']) or 'none'}\n"
        f"{context_block}\n"
        f"Full PUML:\n{puml}\n\n"
        f"{focus_txt} Generate at least {min_cases} test cases.\n"
        f"Return ONLY the JSON array."
    )


# ── Call LLM ────────────────────────────────────────────────────────────────

def call_llm(puml, parsed, knowledge=None):
    payload = {
        "model": LLM_MODEL,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": build_prompt(puml, parsed, knowledge)},
        ],
        "temperature": 0.3,
        "max_tokens":  4096,
        "stream":      False,
    }
    resp = requests.post(f"{LLM_BASE_URL}/chat/completions", json=payload, timeout=LLM_TIMEOUT)
    resp.raise_for_status()
    raw = resp.json()["choices"][0]["message"]["content"].strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw).strip()
    return json.loads(raw)


DIAGRAM_SYSTEM_PROMPT = """You are a senior software architect who writes PlantUML diagrams.
Given a plain-English description of a system, feature, or user flow, generate clean,
valid PlantUML code for the requested diagram type.

Rules:
- Output ONLY valid PlantUML code, starting with @startuml and ending with @enduml.
- No markdown code fences, no explanation, no extra text.
- Use clear actor/participant names derived from the description.
- For sequence diagrams: include realistic messages with HTTP methods/status codes where applicable.
- For use case diagrams: include actors and use cases with proper associations.
- For activity diagrams: include start/stop, decision points, and actions.
- For class diagrams: include relevant attributes and methods.
- Keep it focused and readable — 8-20 lines is ideal."""

def call_llm_diagram(description, diagram_type):
    type_hint = {
        "sequence": "a sequence diagram",
        "usecase":  "a use case diagram",
        "activity": "an activity diagram",
        "class":    "a class diagram",
    }.get(diagram_type, "a sequence diagram")

    payload = {
        "model": LLM_MODEL,
        "messages": [
            {"role": "system", "content": DIAGRAM_SYSTEM_PROMPT},
            {"role": "user", "content": f"Generate {type_hint} for this description:\n\n{description}\n\nReturn ONLY the PlantUML code."},
        ],
        "temperature": 0.4,
        "max_tokens": 1024,
        "stream": False,
    }
    resp = requests.post(f"{LLM_BASE_URL}/chat/completions", json=payload, timeout=LLM_TIMEOUT)
    resp.raise_for_status()
    raw = resp.json()["choices"][0]["message"]["content"].strip()
    raw = re.sub(r"^```(?:plantuml|puml)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw).strip()
    return raw


# ── Fallback (no LLM) ────────────────────────────────────────────────────────

def fallback_generate(parsed):
    cases = []
    title = parsed["title"]
    msgs  = parsed["messages"]
    conds = parsed["conditions"]
    actors_txt = ", ".join(parsed["actors"]) if parsed["actors"] else "required actors"

    def tid(n): return f"TC-{str(n).zfill(3)}"
    i = 1

    # Happy path
    if msgs:
        flow = "\n".join(f"{j+1}. Trigger: {m['from']} \u2192 {m['to']}: {m['label']}" for j,m in enumerate(msgs))
        cases.append({
            "tc_id": tid(i), "module": title, "name": f"Happy Path \u2013 {title}",
            "type": "Positive", "priority": "High",
            "prerequisite": f"System is running. {actors_txt} are available and configured. Test environment is accessible.",
            "steps": flow,
            "expected_output": "All steps complete without error. System responds as designed at each interaction point.",
        }); i+=1

    # Per message
    for m in msgs:
        lbl = m["label"]
        is_create = any(w in lbl.lower() for w in ["create","post","add","register","submit"])
        is_fetch  = any(w in lbl.lower() for w in ["get","fetch","list","read","retrieve"])
        is_auth   = any(w in lbl.lower() for w in ["login","auth","token","sign"])

        if is_auth:
            pre   = f"User account exists in the system. Browser is open on the login page."
            steps = f"1. Open the application login page\n2. Enter valid credentials\n3. Submit the login form\n4. Observe the response"
            exp   = f"System authenticates successfully. Access token or session is returned. User is redirected to the dashboard."
        elif is_create:
            pre   = f"User is authenticated. Required form or API client is open. No duplicate record exists."
            steps = f"1. Navigate to the create {lbl} section\n2. Fill in all required fields with valid data\n3. Submit the request\n4. Note the response"
            exp   = f"System returns 201 Created with the newly created resource object. Record is persisted."
        elif is_fetch:
            pre   = f"User is authenticated. At least one {lbl} record exists in the system."
            steps = f"1. Send GET request for: {lbl}\n2. Observe the HTTP status code\n3. Inspect the response body"
            exp   = f"System returns 200 OK with the requested data in the expected format."
        else:
            pre   = f"System is running. Previous steps have been completed successfully."
            steps = f"1. Initiate: {m['from']} performs '{lbl}'\n2. Observe the response from {m['to']}\n3. Verify the result"
            exp   = f"{m['to']} processes '{lbl}' and responds correctly without error."

        cases.append({
            "tc_id": tid(i), "module": title, "name": f"Verify: {lbl}",
            "type": "Functional", "priority": "Medium" if i > 2 else "High",
            "prerequisite": pre, "steps": steps, "expected_output": exp,
        }); i+=1

    # Alt/condition flows
    for cond in conds:
        cases.append({
            "tc_id": tid(i), "module": title, "name": f"Alt Flow \u2013 {cond}",
            "type": "Negative", "priority": "Medium",
            "prerequisite": f"System is running. Condition can be triggered: {cond}.",
            "steps": f"1. Set up scenario where: {cond}\n2. Perform the triggering action\n3. Observe system behaviour",
            "expected_output": f"System handles '{cond}' gracefully. Appropriate error message or fallback behaviour is shown.",
        }); i+=1

    # Standard edge cases
    edges = [
        ("Invalid / Missing Auth Token", "Negative", "High",
         "System is running. An invalid or expired token is available.",
         "1. Send a request with an invalid Bearer token\n2. Observe the HTTP response code\n3. Read the error body",
         "System returns 401 Unauthorized. Error message is descriptive. No internal data is exposed."),
        ("Empty / Missing Required Fields", "Negative", "High",
         "User is authenticated. The submission form or API endpoint is accessible.",
         "1. Submit a request with all required fields empty\n2. Observe the validation response",
         "System returns 400 Bad Request with a clear list of missing or invalid fields."),
        ("Fetch Non-Existent Resource", "Negative", "Medium",
         "User is authenticated. The target resource ID does not exist in the system.",
         "1. Send a GET request with a non-existent resource ID\n2. Observe the response",
         "System returns 404 Not Found with a descriptive error message."),
        ("Duplicate Record Creation", "Negative", "Medium",
         "User is authenticated. A record with the same unique key already exists.",
         "1. Attempt to create a resource that already exists\n2. Observe the response",
         "System returns 409 Conflict. Duplicate is not created. Error message is meaningful."),
        ("Boundary Value \u2013 Maximum Input Length", "Negative", "Low",
         "User is authenticated. Form or API input field is accessible.",
         "1. Enter a value that exceeds the maximum allowed length\n2. Submit the form\n3. Observe the response",
         "System rejects the input with a 400 error. Appropriate max-length validation message is shown."),
    ]
    for name,typ,pri,pre,steps,exp in edges:
        cases.append({"tc_id":tid(i),"module":title,"name":name,"type":typ,"priority":pri,"prerequisite":pre,"steps":steps,"expected_output":exp}); i+=1

    return cases


# ── Excel Export ─────────────────────────────────────────────────────────────

COLUMNS = [
    ("tc_id",          "TC ID",              12),
    ("module",         "Module",             18),
    ("name",           "Test Case Name",     34),
    ("type",           "Type",               13),
    ("priority",       "Priority",           11),
    ("prerequisite",   "Prerequisite",       40),
    ("steps",          "Test Steps / Process", 50),
    ("expected_output","Expected Output",    40),
]


def build_excel(cases, sheet_title="Test Cases"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    # ── Colour palette (light theme) ──────────────────────────
    HDR_BG    = "FFB8D4F0"   # light blue header background
    HDR_TEXT  = "FF1A2B3C"   # dark navy text on header
    HDR_BORD  = "FF6FA8DC"   # medium blue header border
    ROW_ODD   = "FFFFFFFF"   # white
    ROW_EVEN  = "FFF0F6FF"   # very light blue tint
    ROW_TEXT  = "FF1F2D3D"   # dark text for rows
    BORD_C    = "FFBDD7EE"   # light blue cell border
    TCID_CLR  = "FF1A5EA8"   # blue for TC ID column text

    # Type colours (readable on white)
    TYPE_TEXT = {
        "Positive":   "FF1D7A3A",   # dark green
        "Negative":   "FFC0392B",   # dark red
        "Functional": "FF1A5EA8",   # dark blue
        "Security":   "FF6C3483",   # dark purple
    }
    # Priority colours
    PRI_TEXT = {
        "High":   "FFC0392B",   # dark red
        "Medium": "FFB7770D",   # dark amber
        "Low":    "FF5D6D7E",   # grey
    }

    thin_bord  = Side(style="thin",   color=BORD_C)
    hdr_bord_s = Side(style="medium", color=HDR_BORD)
    cell_border = Border(left=thin_bord, right=thin_bord, top=thin_bord, bottom=thin_bord)
    hdr_border  = Border(left=hdr_bord_s, right=hdr_bord_s, top=hdr_bord_s, bottom=hdr_bord_s)

    # ── Header row ────────────────────────────────────────────
    for ci, (_, label, _) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font      = Font(name="Calibri", bold=True, color=HDR_TEXT, size=11)
        c.fill      = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = hdr_border
    ws.row_dimensions[1].height = 32

    # ── Data rows ─────────────────────────────────────────────
    for ri, tc in enumerate(cases, 2):
        fill_clr = ROW_ODD if ri % 2 == 0 else ROW_EVEN
        for ci, (key, _, _) in enumerate(COLUMNS, 1):
            val = tc.get(key, "")
            c   = ws.cell(row=ri, column=ci, value=val)
            c.fill      = PatternFill("solid", fgColor=fill_clr)
            c.alignment = Alignment(
                vertical="top", wrap_text=True,
                horizontal="center" if key in ("tc_id", "type", "priority") else "left"
            )
            c.border = cell_border

            if key == "tc_id":
                c.font = Font(name="Calibri", bold=True, color=TCID_CLR, size=10)
            elif key == "type":
                clr = TYPE_TEXT.get(val, "FF1A5EA8")
                c.font = Font(name="Calibri", bold=True, color=clr, size=10)
            elif key == "priority":
                clr = PRI_TEXT.get(val, "FF5D6D7E")
                c.font = Font(name="Calibri", bold=True, color=clr, size=10)
            else:
                c.font = Font(name="Calibri", color=ROW_TEXT, size=10)

        ws.row_dimensions[ri].height = 60

    # ── Column widths & freeze ────────────────────────────────
    for ci, (_, _, w) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Flask Routes ──────────────────────────────────────────────────────────────

@app.route("/health")
def health():
    llm_ok, llm_err = False, None
    try:
        r = requests.get(f"{LLM_BASE_URL}/models", timeout=5)
        llm_ok = r.status_code == 200
    except Exception as e:
        llm_err = str(e)
    return jsonify({
        "server": "ok",
        "llm_url": LLM_BASE_URL,
        "llm_model": LLM_MODEL,
        "llm_reachable": llm_ok,
        "llm_error": llm_err,
    })


@app.route("/generate", methods=["POST"])
def generate():
    body = request.get_json(force=True, silent=True) or {}
    puml = body.get("puml","").strip()
    knowledge = body.get("knowledge", {}) or {}
    if not puml:
        return jsonify({"error":"No PUML content provided"}), 400

    parsed = parse_puml(puml)
    try:
        cases = call_llm(puml, parsed, knowledge)
        source = "llm"
    except Exception as e:
        app.logger.error(f"LLM failed: {e}")
        return jsonify({"error": f"LLM generation failed: {str(e)}. Make sure your LLM is running and reachable at {LLM_BASE_URL}"}), 500

    # ── Validate: filter out any non-dict entries (LLM sometimes returns
    #    error strings wrapped in a JSON array instead of TC objects) ──
    valid_cases = []
    for item in cases:
        if not isinstance(item, dict):
            app.logger.warning(f"Non-dict item in LLM response filtered out: {str(item)[:100]}")
            continue
        # If the item looks like an embedded error message, reject the whole batch
        name_val = str(item.get("name", "") or "").lower()
        steps_val = str(item.get("steps", "") or "").lower()
        if any(phrase in name_val or phrase in steps_val for phrase in [
            "llm generation failed", "error:", "failed to", "unable to", "cannot generate"
        ]):
            app.logger.warning(f"LLM error string detected in case fields: {item}")
            continue
        valid_cases.append(item)

    if not valid_cases:
        return jsonify({
            "error": "LLM returned no valid test cases. The model may have returned an error message instead of JSON. Try again or check your LLM logs."
        }), 500

    # Guarantee required keys on all valid cases
    required = ["tc_id","module","name","type","priority","prerequisite","steps","expected_output"]
    for i, tc in enumerate(valid_cases):
        tc.setdefault("tc_id", f"TC-{str(i+1).zfill(3)}")
        tc.setdefault("status", "Pending")
        for k in required:
            tc.setdefault(k, "")

    return jsonify({"cases": valid_cases, "source": source, "title": parsed["title"], "llm_model": LLM_MODEL})


@app.route("/generate-diagram", methods=["POST"])
def generate_diagram():
    body = request.get_json(force=True, silent=True) or {}
    description = body.get("description", "").strip()
    diagram_type = body.get("diagram_type", "sequence")
    if not description:
        return jsonify({"error": "No description provided"}), 400
    try:
        puml = call_llm_diagram(description, diagram_type)
        return jsonify({"puml": puml, "llm_model": LLM_MODEL})
    except Exception as e:
        app.logger.error(f"Diagram LLM failed: {e}")
        return jsonify({"error": f"Diagram generation failed: {str(e)}. Make sure your LLM is running and reachable at {LLM_BASE_URL}"}), 500


@app.route("/render-status", methods=["GET"])
def render_status():
    """Tell the frontend whether local plantuml.jar rendering is available."""
    return jsonify({
        "jar_available": plantuml_jar_available(),
        "jar_path": PLANTUML_JAR,
        "renderer": "plantuml.jar (local)" if plantuml_jar_available() else "kroki.io (online fallback)",
    })


@app.route("/render-diagram", methods=["POST"])
def render_diagram():
    """
    Render PUML text to an image (PNG or SVG).
    Body: { "puml": "@startuml...", "format": "png" }
    Returns: image bytes with appropriate mimetype
    """
    body = request.get_json(force=True, silent=True) or {}
    puml = body.get("puml", "").strip()
    fmt  = body.get("format", "png").lower()

    if not puml:
        return jsonify({"error": "No PUML content provided"}), 400
    if "@startuml" not in puml:
        puml = "@startuml\n" + puml + "\n@enduml"
    if fmt not in ("png", "svg"):
        fmt = "png"

    try:
        image_bytes = render_puml(puml, fmt)
        mimetype = "image/svg+xml" if fmt == "svg" else "image/png"
        return send_file(BytesIO(image_bytes), mimetype=mimetype)
    except subprocess.TimeoutExpired:
        return jsonify({"error": "Rendering timed out after 30s. Diagram may be too complex."}), 500
    except Exception as e:
        app.logger.error(f"Render failed: {e}")
        renderer = "plantuml.jar" if plantuml_jar_available() else "Kroki.io"
        return jsonify({
            "error": f"Rendering failed via {renderer}: {str(e)}"
        }), 500


@app.route("/export", methods=["POST"])
def export_excel():
    body     = request.get_json(force=True, silent=True) or {}
    cases    = body.get("cases", [])
    filename = body.get("filename", "testcases.xlsx")
    title    = body.get("title",    "Test Cases")
    if not cases:
        return jsonify({"error":"No cases"}), 400
    try:
        buf = build_excel(cases, sheet_title=title[:31])
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500



@app.route("/extract-diagram", methods=["POST"])
def extract_diagram():
    """
    Extract PlantUML diagram text from an image using LLM vision.
    Body: { "image_base64": "...", "image_mime": "image/png" }
    Returns: { "puml": "@startuml\n..." }
    """
    body      = request.get_json(force=True, silent=True) or {}
    image_b64 = body.get("image_base64", "").strip()
    image_mime= body.get("image_mime", "image/png")

    if not image_b64:
        return jsonify({"error": "No image_base64 provided"}), 400

    prompt = """You are a PlantUML expert. Analyse this sequence diagram image carefully.
Extract ALL information visible in the image and reconstruct it as valid PlantUML code.

Include:
- @startuml / @enduml
- title (if visible)
- All actors/participants
- All arrows and messages with exact labels
- Any alt/opt/loop blocks with their conditions
- Any notes

Return ONLY the PlantUML code. No explanation, no markdown fences."""

    try:
        payload = {
            "model": LLM_MODEL,
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{image_mime};base64,{image_b64}"
                            }
                        },
                        {
                            "type": "text",
                            "text": prompt
                        }
                    ]
                }
            ],
            "max_tokens": 2048,
            "temperature": 0.1,
            "stream": False,
        }
        resp = requests.post(
            f"{LLM_BASE_URL}/chat/completions",
            json=payload,
            timeout=LLM_TIMEOUT,
        )
        resp.raise_for_status()
        raw = resp.json()["choices"][0]["message"]["content"].strip()

        # Strip markdown fences if LLM added them
        raw = re.sub(r"^```(?:plantuml|puml|uml)?\s*", "", raw, flags=re.IGNORECASE)
        raw = re.sub(r"\s*```$", "", raw).strip()

        # Ensure it looks like PUML
        if "@startuml" not in raw:
            raw = "@startuml\n" + raw + "\n@enduml"

        return jsonify({"puml": raw})

    except Exception as e:
        app.logger.error(f"Image extraction failed: {e}")
        return jsonify({
            "error": f"Image extraction failed: {str(e)}. Make sure your LLM supports vision (e.g. llava, gpt-4o)."
        }), 500


@app.route("/compare-diagrams", methods=["POST"])
def compare_diagrams():
    """
    Compare two PUML diagrams and return AI explanation of changes.
    Body: { "old_puml": "...", "new_puml": "...", "changes": [...] }
    """
    body = request.get_json(force=True, silent=True) or {}
    old_puml = body.get("old_puml", "").strip()
    new_puml = body.get("new_puml", "").strip()
    changes  = body.get("changes", [])

    if not old_puml or not new_puml:
        return jsonify({"error": "Both old_puml and new_puml are required"}), 400

    change_list = "\n".join([
        f"- [{c.get('type','').upper()}] {c.get('category','')}: {c.get('description','')}"
        for c in changes
    ]) if changes else "No structural changes detected."

    prompt = f"""You are a senior QA engineer reviewing two versions of a PlantUML sequence diagram.

OLD diagram:
{old_puml[:800]}

NEW diagram:
{new_puml[:800]}

Detected changes:
{change_list}

Write a clear, concise human-readable explanation (3-4 paragraphs, under 200 words) covering:
1. What structurally changed between the old and new diagram
2. Why these changes likely matter from a QA/testing perspective
3. Which existing test cases may need to be updated or which new ones should be created

Return ONLY the explanation text, no markdown, no bullet points."""

    try:
        payload = {
            "model": LLM_MODEL,
            "messages": [
                {"role": "system", "content": "You are a senior QA engineer who explains software diagram changes clearly."},
                {"role": "user", "content": prompt},
            ],
            "temperature": 0.4,
            "max_tokens": 400,
            "stream": False,
        }
        resp = requests.post(f"{LLM_BASE_URL}/chat/completions", json=payload, timeout=LLM_TIMEOUT)
        resp.raise_for_status()
        explanation = resp.json()["choices"][0]["message"]["content"].strip()
        return jsonify({"explanation": explanation})
    except Exception as e:
        app.logger.error(f"Compare LLM failed: {e}")
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    print("="*55)
    print("  PUML TestCase Generator")
    print("="*55)
    print(f"  LLM  : {LLM_BASE_URL}  [{LLM_MODEL}]")
    print()
    print("  >>> Open in browser:  http://localhost:5050")
    print()
    print("  Both files must be in the same folder:")
    print("    backend.py")
    print("    index.html")
    print("="*55)
    app.run(host="0.0.0.0", port=5050, debug=False)
